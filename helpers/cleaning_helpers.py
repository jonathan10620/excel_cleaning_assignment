from re import L
from openpyxl import load_workbook
from pprint import pprint
import csv
from openpyxl.styles import PatternFill
from .style_helpers import red_fill
from datetime import datetime
import collections

import datedelta

wb = load_workbook("main.xlsx")
ws = wb.active


def get_header_names():
    # get header names from spreadsheet returned as list
    try:
        header_row = ws[1]
    except Exception as e:
        print(e)
        input("issue with getting header row, press enter to exit")
        return None

    # create list of header names
    header_names = []
    for cell in header_row:
        header_names.append(cell.value)

    return header_names


def return_missing_values(col=None):
    """
    returns a list of coordinates with missing values.
    if col is specified, returns only missing values in that column
    """
    if col is not None:
        column = (col, col)
    else:
        column = (1, ws.max_column)

    # reads in spreadsheet and returns list of coordinate tuples for cells with missing values
    missing_coordinates = []

    for row in ws.iter_cols(
        min_row=2, max_row=ws.max_row, min_col=column[0], max_col=column[1]
    ):
        for cell in row:
            if cell.value is None:
                missing_coordinates.append((cell.column, cell.row))

    return missing_coordinates

def identify_duplicates(l):
    return [item for item, count in collections.Counter(l).items() if count > 1]



def return_duplicate_values_within_column(col):
    """
    returns a list of coordinates with duplicate values.
    col must be specified.
    """
    # reads in spreadsheet and returns list of coordinate tuples for cells with duplicate values
    duplicate_coordinates = []
    cell_values = []
    for row in ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=col, max_col=col):
        for cell in row:
            if cell.value is not None:
                if cell.value in cell_values:
                    duplicate_coordinates.append((cell.column, cell.row))
                else:
                    cell_values.append(cell.value)

    return (duplicate_coordinates, identify_duplicates(read_col(col)))


def return_full_names():
    """
    returns a list of full names for all merged cells
    """
    full_names = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=4):
        full_name = []
        for cell in row:
            if cell.value is not None:
                full_name.append(cell.value)

        full_names.append(f"{full_name[0]} {full_name[1]}. {full_name[2]}")

    return full_names


def read_col(col, row_id=False):
    """
    returns a list of values in a column
    accepts header name, or integer as column
    """
    if not isinstance(col, int):
        if col not in get_header_names():
            print("column name not found")
            raise Exception("column name not found")
        col_index = None
        for n, column in enumerate(get_header_names(), 1):
            if col == column:
                col_index = n
    else:
        col_index = col

    if row_id:
        # create dictionary of values
        return {
            n: val[0].value
            for n, val in enumerate(
                ws.iter_rows(min_col=col_index, max_col=col_index, min_row=2), 1
            )
        }
        pass
    else:
        values = []
        for row in ws.iter_rows(min_col=col_index, max_col=col_index, min_row=2):
            values.append(row[0].value)

    return values


def read_row(row, headers=False):
    """
    returns a list of values in a row
    accepts header name, or integer as row
    """
    if not isinstance(row, int):
        if row not in get_header_names():
            print("row name not found")
            raise Exception("row name not found")
        row_index = None
        for n, row in enumerate(get_header_names(), 1):
            if row == row:
                row_index = n
    else:
        row_index = row + 1

    if headers:
        # create dictionary of values
        return {
            head: val[0].value
            for head, val in zip(
                get_header_names(),
                ws.iter_cols(
                    min_row=row_index,
                    max_row=row_index,
                    min_col=1,
                    max_col=ws.max_column,
                ),
            )
        }
    else:
        values = []
        for cell in ws.iter_cols(
            min_row=row_index, max_row=row_index, min_col=1, max_col=ws.max_column
        ):
            values.append(cell[0].value)

        return values


def compare_rows(row_1, row_2):
    same = []
    different = []

    for h in zip(get_header_names(), read_row(row_1), read_row(row_2)):
        if h[1] == h[2] and h[1] is not None:
            same.append({h[0]: h[1]})
        elif h[1] is not None:
            different.append({h[0]: h[1]})

    with open(f"findings/findings{row_1}_{row_2}.txt", "w+") as f:
        f.write("SAME:\n")
        for s in same:
            f.write(f"{s}\n")

        f.write("\nDIFFERENT:\n")
        for d in different:
            f.write(f"{d}\n")


def read_col_range(col1, col2):
    final_list = []
    for row in ws.iter_rows(min_col=col1, max_col=col2, min_row=2, max_row=ws.max_row):
        row_list = []
        for cell in row:
            if cell.value is not None:
                row_list.append(cell.value)
        final_list.append(row_list)

    print(final_list)

    return final_list


def read_zip_city_county():
    with open("dictionary/tx_zips.csv") as read_obj:
        # pass the file object to reader() to get the reader object
        reader = csv.reader(read_obj)
        data = list(reader)
        zip_dict = {
            row[0]: {"city": row[1], "county": row[2].replace("County", "").strip()}
            for row in data
        }

    # for each row, read in zip, city, county
    # H is zip
    # F is city
    # J is county
    data_zip_dict = {}
    city_errors = []
    county_errors = []
    for row in zip(ws["H"], ws["F"], ws["J"]):
        try:
            data_zip_dict[str(row[0].value)] = {
                "city": row[1].value.strip(),
                "county": row[2].value.strip(),
            }
        except AttributeError:
            input("found none at row {}".format(row[0].row))

        x_zip_code = str(row[0].value)
        if x_zip_code in zip_dict:
            if data_zip_dict[x_zip_code]["city"] != zip_dict[x_zip_code]["city"]:
                print("incorrect city")
                city_errors.append(row[1].coordinate)
                print(
                    f'{x_zip_code} {data_zip_dict[x_zip_code]["city"]}:{zip_dict[x_zip_code]["city"]}'
                )
                row[1].value = zip_dict[x_zip_code]["city"]
                wb.save("main.xlsx")

            elif data_zip_dict[x_zip_code]["county"] != zip_dict[x_zip_code]["county"]:
                print("incorrect county")
                county_errors.append(row[2].coordinate)
                print(
                    f'{x_zip_code} {data_zip_dict[x_zip_code]["county"]}:{zip_dict[x_zip_code]["county"]}'
                )
                row[2].value = zip_dict[x_zip_code]["county"]
                wb.save("main.xlsx")


def death_code():
    # with open('dictionary/dc_status.csv') as read_obj:
    #     # pass the file object to reader() to get the reader object
    #     reader = csv.reader(read_obj)
    #     data = list(reader)
    #     death_codes = {row[0]: row[1] for row in data}

    for status, death in zip(ws["K"], ws["L"]):
        if str(status.value) in ["20", "40", "48", "41", "42"]:
            death.value = "Y"
        else:
            death.value = "N"

    wb.save("main.xlsx")


def newborn():
    for age, newborn in zip(ws["W"], ws["M"]):
        if age.value == 0:
            newborn.value = "Y"
        else:
            newborn.value = "N"
    wb.save("main.xlsx")


def ethnicity():
    for eth_code, desc in zip(ws["P"], ws["Q"]):

        if eth_code.value == 1:
            desc.value = "Hispanic"
        elif eth_code.value == 2:
            desc.value = "Non-Hispanic"
        else:
            eth_code.fill = PatternFill(start_color="FF7E79", fill_type="solid")
            desc.fill = PatternFill(start_color="FF7E79", fill_type="solid")

    wb.save("main.xlsx")


def salutation():
    for sex, sal, mar in zip(ws["N"], ws["S"], ws["R"]):
        if sex.value == "M":
            sal.value = "Mr."
        elif sex.value == "F" and mar.value == "Married":
            sal.value = "Mrs."
        elif sex.value == "F":
            sal.value = "Ms."
        else:
            sal.value = "N/A"

    wb.save("main.xlsx")


def los():
    for start, end, los in zip(ws["T"], ws["U"], ws["V"]):
        if isinstance(start.value, str):
            continue

        los_actual_date = end.value - start.value
        los_actual = str(los_actual_date.days)
        print(los_actual)

        if los.value != los_actual:
            los.value = los_actual
        else:
            los.value = los_actual

    wb.save("main.xlsx")


def insurance():
    with open("dictionary/insurance.csv", "r") as fh:
        reader = csv.reader(fh)
        insurance_dict = {row[0].strip(): row[1].strip() for row in reader}

    print(insurance_dict)
    for insurance, src_code in zip(ws["Y"], ws["X"]):

        if src_code.value in insurance_dict:
            insurance.value = insurance_dict[src_code.value]
        else:
            insurance.value = "N/A"
            src_code.fill = PatternFill(start_color="FF7E79", fill_type="solid")

    wb.save("main.xlsx")


def hr():
    anomalies = []
    for new, hr in zip(ws["M"], ws["BA"]):
        # if newborn
        if new.value == "Y":
            try:
                if hr.value < 100:
                    hr.fill = red_fill
                    anomalies.append(hr.coordinate)
            except Exception as e:
                print("Empty cell found")
                hr.fill = red_fill
        else:
            try:
                if hr.value > 200 or hr.value < 30:
                    hr.fill = red_fill
                    anomalies.append(hr.coordinate)
            except Exception as e:
                print("Empty cell found")
                hr.fill = red_fill

    print(len(anomalies))
    # wb.save('main.xlsx')


def bmi_desc():
    for n, cell in enumerate(ws["AY"]):
        if cell.value == "BMI":
            continue
        elif cell.value < 18.5:
            ws["AZ" + str(n + 1)].value = "Underweight"
        elif cell.value >= 18.5 and cell.value <= 24.9:
            ws["AZ" + str(n + 1)].value = "Normal"
        elif cell.value >= 25 and cell.value <= 29.9:
            ws["AZ" + str(n + 1)].value = "Overweight"
        elif cell.value >= 30:
            ws["AZ" + str(n + 1)].value = "Obese"


def convert_height_to_inches(string):
    split = string.strip().split(" ")

    return int(split[0].rstrip("''")) * 12 + int(split[1].rstrip('"'))


def clean_height():
    incorrect_format = []
    incorrect_conversion = []
    for height, height_inches in zip(ws["AT"], ws["AU"]):
        if isinstance(height.value, str):
            continue
        if isinstance(height.value, float):
            print("heigh in incorrect format")
            height.fill = red_fill
            incorrect_format.append(height.coordinate)
        elif height_inches.value != convert_height_to_inches(height.value):
            print("height in incorrect conversion")
            height_inches.value = convert_height_to_inches(height.value)
            height_inches.fill = red_fill
            incorrect_conversion.append(height.coordinate)
        else:
            print("empty cell found")
            height.fill = red_fill
    wb.save("main.xlsx")


def convert_weight():
    for lb, kg in zip(ws["AV"], ws["AX"]):
        if isinstance(lb.value, str):
            continue
        try:
            kg.value = lb.value * 0.453592
        except Exception as e:
            print("empty cell found")
            kg.fill = red_fill
    wb.save("main.xlsx")


def bmi_descr(value):
    if value < 18.5:
        return "Underweight"
    elif value >= 18.5 and value <= 24.9:
        return "Normal"
    elif value >= 25 and value <= 29.9:
        return "Overweight"
    elif value >= 30:
        return "Obese"


def inches():
    for inches in ws["AU"]:
        try:
            if int(inches.value) == 0:
                inches.value = "N/A"
                inches.fill = red_fill
        except Exception as e:
            print("we passed")
            pass
    wb.save("main.xlsx")


def bmi_fix():
    for lb, inches, bmi, bmi_desc in zip(ws["AV"], ws["AU"], ws["AY"], ws["AZ"]):
        if isinstance(lb.value, str) and not lb.value.isdigit():
            continue
        try:
            bmi.value = int(
                int(lb.value) * 0.453592 / (int(inches.value) * 0.0254) ** 2
            )
            bmi_desc.value = bmi_descr(bmi.value)
        except Exception as e:
            bmi.value = "N/A"
            bmi.desc = "N/A"
            bmi.fill = red_fill

    wb.save("main.xlsx")


def fix_none():
    for row in ws.iter_rows(
        max_row=ws.max_row,
        min_row=ws.min_row,
        max_col=ws.max_column,
        min_col=ws.min_column,
    ):
        print(row)
        for cell in row:
            print(cell)
            if cell.value is None or cell.value == "":
                cell.value = "N/A"
                cell.fill = red_fill
    wb.save("main.xlsx")


def daily_charge():
    for total, los, daily in zip(ws["Z"], ws["V"], ws["AC"]):
        if daily.value == "Daily_Chrg":
            continue

        try:
            daily.value = float("{0:.2f}".format(int(total.value) / int(los.value)))
        except Exception as e:
            print(e)
            print(total.value)
            print(los.value)
            input("error found?")

    wb.save("main.xlsx")


def cleanse_equations(column):
    for cell in ws[column]:
        try:
            print(str(cell.value))
            if str(cell.value[0]) == "=":
                cell.value = "N/A"
                cell.fill = red_fill
        except Exception as e:
            pass
    wb.save("main.xlsx")


def illness():
    code_dic = {1: "Acute", 2: "Sub-Acute", 3: "Moderate", 4: "Managed"}

    for code, desc in zip(ws["AN"], ws["AO"]):
        try:
            if int(code.value) in code_dic:
                desc.value = code_dic[int(code.value)]

        except ValueError:
            print("encountered a non integer")
            pass
    wb.save("main.xlsx")
