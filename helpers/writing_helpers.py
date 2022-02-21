from openpyxl import load_workbook, Workbook

from openpyxl.styles import PatternFill
from helpers.cleaning_helpers import *
from helpers.style_helpers import *


excel_file_read = "main.xlsx"
wb_read = load_workbook(excel_file_read)
ws_read = wb_read.active

excel_file_write = "main_cleaned.xlsx"
wb_write = Workbook()
ws = wb_write.active


def update_cell(col, row, value):
    """
    returns cell object at row, col
    """
    if not isinstance(col, int):
        if col not in get_header_names():
            print("column name not found")
            raise Exception("column name not found")
        col_index = None
        for n, column in enumerate(get_header_names(), 1):
            if col == column:
                col_index = n
    else:
        col_index = col

    ws.cell(row, col_index).value = value
    wb_write.save(excel_file_write)


def write_col_headers():
    # writes header list to spreadsheet
    # necessary for writing headers and must be run before writing data
    for n, header in enumerate(get_header_names(), 1):
        ws.cell(1, n).value = header
        ws.cell(1, n).fill = PatternFill(start_color="00FA92", fill_type="solid")

    wb_write.save(excel_file_write)


def find_all_missing_rows(values):
    l = []
    missing_indexes = []
    for n, row in enumerate(values, 1):
        if row == None:
            missing_indexes.append(n)
            print("missing entry at row: " + str(n))
        l.append(row)

    return missing_indexes


def get_col_values_by_row(row_index):
    # returns dict of header:value pairs for row_index
    col_values = {
        header: val[0].value
        for header, val in zip(
            get_header_names(),
            ws_read.iter_cols(min_row=row_index + 1, max_row=row_index + 1),
        )
    }

    return col_values


def check_for_missing_values(col):
    # iterates through each row in column and checks for missing values
    # if missing values are found, prompts user to fill in missing values either by one, or all
    if not isinstance(col, int):
        if col not in get_header_names():
            print("column name not found")
            raise Exception("column name not found")
        col_index = None
        for n, column in enumerate(get_header_names(), 1):
            if col == column:
                col_index = n
    else:
        col_index = col

    for row in read_col(col_index):
        if row == None:
            print(f"missing value in cell at {row}")
            info = input("would you like to read row? (y/n)")
            if info == "y":
                # crazy way to get list of all values in row
                col_values = get_col_values_by_row(row[0].row, ws)
                print(col_values)

            change_one = input(
                f"would you like to give a one-time value to cell at {row[0].coordinate}? (y/n)"
            )

            if change_one == "y":
                value = input("what value would you like to give?: ")
                row[0].value = value

                # wb.save(excel_file)
                continue

            elif change_one == "n":
                _all = input(
                    f"would you like to change all missing values in {col}? (y/n) "
                )
                if _all == "y":
                    # change all missing values in column
                    missing_data = find_all_missing_rows(read_col(col))
                    print(f"missing data found at rows: {missing_data}")

                    all_val = input(
                        f"what value would you like to give to all empty cells in col {col}?: "
                    )
                    for cell in missing_data:
                        print("changing cell at row " + str(cell))
                        update_cell(col, ws, cell, all_val)

                    # wb.save(excel_file)
                    print("all missing values in column have been changed")
                    print("exiting now")
                    return
                else:
                    print("no changes made. exiting")
                    return

    return False


def write_data(col):
    for row in ws_read.iter_rows(min_row=2, min_col=col, max_col=col):
        update_cell(col, row[0].row, row[0].value)

    wb_write.save(excel_file_write)


# TODO: add function to update remove and values in column
def duplicate_row_list(values):
    # iterates through list of values and checks for duplicates and if so returns list of index where duplicates are found
    l = []
    duplicate_indexes = []
    for n, row in enumerate(values, 2):
        if row in l:
            duplicate_indexes.append((n, row))
        l.append(row)

    return duplicate_indexes


def remove_all_duplicates_from_column(col, new_value=None):
    # iterates through each row in column and checks for duplicates
    # if duplicates are found, prompts user to remove duplicates
    if col not in get_header_names():
        print("column name not found")
        raise Exception("column name not found")

    col_index = None
    for n, col in enumerate(get_header_names(), 1):
        if col == col:
            col_index = n

    l = duplicate_row_list(read_col(col))

    for row in ws.iter_rows(min_col=col_index, max_col=col_index, min_row=2):
        # writing to new documennt each row of column
        # if header just write with blue fill

        ws.cell(row[0].row, col_index).value = row[0].value

        # if duplicate is found, prompt user to remove duplicate
        if (row[0].row, row[0].value) in l:
            print(
                f"duplicate found at row {row[0].row} with value {row[0].value} and coordinate {row[0].coordinate}"
            )
            info = input("would you like to remove cell value? (y/n)")
            if info == "y":
                ws.cell(row[0].row, col_index).value = new_value
                ws.cell(row[0].row, col_index).fill = red_fill
                continue
            else:
                print("no changes made. exiting")
                return

        wb_write.save(excel_file_write)
